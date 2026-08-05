import openpyxl
import json
import urllib.request
import os
import sys

SUPABASE_URL = os.environ["NEXT_PUBLIC_SUPABASE_URL"]
SUPABASE_KEY = os.environ["SUPABASE_SECRET_KEY"]

def supabase_post(path, body):
    url = f"{SUPABASE_URL}/rest/v1/{path}?on_conflict=date,gold_type_id"
    req = urllib.request.Request(url, data=json.dumps(body).encode(), method="POST")
    req.add_header("apikey", SUPABASE_KEY)
    req.add_header("Authorization", f"Bearer {SUPABASE_KEY}")
    req.add_header("Content-Type", "application/json")
    req.add_header("Prefer", "resolution=merge-duplicates")
    try:
        with urllib.request.urlopen(req) as resp:
            return resp.status, None
    except urllib.error.HTTPError as e:
        return e.code, e.read().decode()

# Mapping Excel columns to gold_type_ids
LM_IDS = ["antam-0.5","antam-1","antam-2","antam-3","antam-5","antam-10","antam-25","antam-50","antam-100"]
BB_LM_IDS = ["bb-certi-1-2","bb-certi-3-5","bb-certi-10-25","bb-certi-50-100","bb-non-rm","bb-retro","bb-merek-lain"]
BB_PH_IDS = ["ph-k24s","ph-k24","ph-k23","ph-k22","ph-k21","ph-k20","ph-k19","ph-k18","ph-k17","ph-k16","ph-k15","ph-k14","ph-k13","ph-k12","ph-k11","ph-k10","ph-k9","ph-k8","ph-k7","ph-k6"]
LL_IDS = {"Palladium /gr": "ll-palladium", "Perak /gr": "ll-perak"}

wb = openpyxl.load_workbook('Historis_Pricelist_15Jul-4Aug2026.xlsx')

rows_to_insert = []

# === Harga Jual LM ===
ws_lm = wb['Harga Jual LM']
for row in ws_lm.iter_rows(min_row=5, max_row=ws_lm.max_row, values_only=True):
    if row[1] != 'Update' or row[2] is None:
        continue
    date = str(row[0])[:10]
    dasar = int(row[11])
    for i, gid in enumerate(LM_IDS):
        if row[2+i] is not None:
            price = int(row[2+i])
            if gid in ("antam-50", "antam-100"):
                buy_price = price  # total price for 50/100gr
            else:
                buy_price = price  # total price per product
            rows_to_insert.append({
                "date": date,
                "gold_type_id": gid,
                "buy_price": buy_price,
                "sell_price": 0,
                "base_price": dasar,
            })

# === Buyback LM ===
ws_bb = wb['Buyback LM']
for row in ws_bb.iter_rows(min_row=5, max_row=ws_bb.max_row, values_only=True):
    if row[1] != 'Update' or row[2] is None:
        continue
    date = str(row[0])[:10]
    acuan = int(row[2])
    for i, gid in enumerate(BB_LM_IDS):
        if row[2+i] is not None:
            rows_to_insert.append({
                "date": date,
                "gold_type_id": gid,
                "buy_price": 0,
                "sell_price": int(row[2+i]),
                "base_price": acuan,
            })

# === Buyback Perhiasan ===
ws_ph = wb['Buyback Perhiasan']
for row in ws_ph.iter_rows(min_row=5, max_row=ws_ph.max_row, values_only=True):
    if row[1] != 'Update' or row[2] is None:
        continue
    date = str(row[0])[:10]
    for i, gid in enumerate(BB_PH_IDS):
        if row[2+i] is not None:
            rows_to_insert.append({
                "date": date,
                "gold_type_id": gid,
                "buy_price": 0,
                "sell_price": int(row[2+i]),
                "base_price": 0,
            })

# === Logam Lain ===
ws_ll = wb['Logam Lain']
for row in ws_ll.iter_rows(min_row=5, max_row=ws_ll.max_row, values_only=True):
    if row[1] != 'Update':
        continue
    date = str(row[0])[:10]
    for col_idx in range(2, 4):
        col_name = ws_ll.cell(row=4, column=col_idx+1).value
        gid = LL_IDS.get(col_name)
        if gid and row[col_idx] is not None:
            rows_to_insert.append({
                "date": date,
                "gold_type_id": gid,
                "buy_price": 0,
                "sell_price": int(row[col_idx]),
                "base_price": 0,
            })

print(f"Total rows to insert: {len(rows_to_insert)}")
print()

# Insert in batches of 100
batch_size = 50
total = len(rows_to_insert)
for i in range(0, total, batch_size):
    batch = rows_to_insert[i:i+batch_size]
    status, error = supabase_post("price_history", batch)
    if status in (200, 201):
        print(f"  Batch {i//batch_size+1}/{(total+batch_size-1)//batch_size}: OK ({len(batch)} rows)")
    else:
        print(f"  Batch {i//batch_size+1}: ERROR {status} - {error}")
        break

print()
print("Import selesai!")
